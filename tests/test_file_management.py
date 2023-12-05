import os
from io import BytesIO

import openpyxl
import pandas as pd
import pytest
from flask import Flask
from werkzeug.datastructures import FileStorage

from app.modules.file_management import FileManagement


@pytest.fixture
def app():
    app = Flask(__name__)
    app.config['DATA_RAW_DIR'] = 'data/raw'
    app.config['DATA_CLEAN_DIR'] = 'data/clean'
    app.config['DATA_PROCESSED_DIR'] = 'data/processed'
    app.config['ALLOWED_EXTENSIONS'] = {'xls', 'xlsx', 'csv'}
    return app


@pytest.fixture
def file_manager(app):
    with app.app_context():
        yield FileManagement()


def test_list_files_raw(file_manager, app):
    with app.app_context():
        listed_files = file_manager.list_files('raw')
        all_files_allowed = all(
            file.split('.')[-1] in app.config['ALLOWED_EXTENSIONS']
            for file in listed_files
        )
        assert all_files_allowed, "Some listed files in raw directory have disallowed extensions"


def test_upload_file_no_overwrite(file_manager, app):
    with app.app_context():
        # First upload
        dummy_file = FileStorage(stream=BytesIO(b"dummy file content"), filename="dummy.xlsx")
        message = file_manager.upload_file(dummy_file, 'raw')
        assert message == 'File dummy.xlsx uploaded successfully to raw directory'

        # Attempt to upload the same file again
        second_message = file_manager.upload_file(dummy_file, 'raw')
        assert second_message == 'File dummy.xlsx already exists in raw directory', "File overwrite check failed"


def test_delete_file(file_manager, app):
    # Setup: Ensure a test file exists
    test_file_path = os.path.join(app.config['DATA_RAW_DIR'], 'delete_test.csv')
    with open(test_file_path, 'w') as file:
        file.write('test content')

    # Test deleting the file
    delete_message = file_manager.delete_file('delete_test.csv', 'raw')
    assert delete_message == 'File delete_test.csv deleted successfully'
    assert not os.path.exists(test_file_path), "File was not deleted"


def test_list_worksheets(file_manager, app, tmp_path):
    # Create a test .xlsx file with two worksheets
    test_file_path = tmp_path / "test_workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.create_sheet(title="Sheet1")
    wb.create_sheet(title="Sheet2")
    wb.save(test_file_path)

    # Convert test_file_path to string
    test_file_path_str = str(test_file_path)

    # Test listing worksheets
    worksheets = file_manager.list_worksheets(test_file_path_str)
    assert len(worksheets) == 3
    assert "Sheet" in worksheets and "Sheet1" in worksheets and "Sheet2" in worksheets, ("Worksheets names not listed "
                                                                                         "correctly")


def test_view_csv_contents(file_manager, app, tmp_path):
    # Create a test CSV file
    test_csv_path = tmp_path / "test.csv"
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    df.to_csv(test_csv_path, index=False)

    # Test viewing CSV content
    content_html = file_manager.view_csv_contents(str(test_csv_path), ['A'])
    assert '<th>A</th>' in content_html
    assert '<td>1</td>' in content_html  # Checking if first row's data is present


def test_view_spreadsheet_contents(file_manager, app, tmp_path):
    # Create a test .xlsx file
    test_xlsx_path = tmp_path / "test.xlsx"
    df = pd.DataFrame({'A': [1, None, 3], 'B': [4, 5, None]})
    df.to_excel(test_xlsx_path, index=False)

    # Test viewing spreadsheet content
    content, metadata = file_manager.view_spreadsheet_contents(str(test_xlsx_path), 'Sheet1')
    assert metadata['total_rows'] == 3
    assert 'A' in content and 'B' in content
    assert content['A']['column_name'] == 'A'
    assert content['B']['column_name'] == 'B'
    assert content['A']['first_item'] == 1
    assert content['A']['missing_count'] == 1  # One missing value in column A


def test_save_as_csv(file_manager, app, tmp_path):
    # Create a test DataFrame
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})

    # Test saving as CSV
    save_path = tmp_path / "saved.csv"
    column_selection = {'A': 'A', 'B': 'B_renamed'}  # Renaming column 'B' to 'B_renamed'
    file_manager.save_as_csv(df, column_selection, str(save_path))

    # Verify the saved file
    saved_df = pd.read_csv(save_path)
    assert 'A' in saved_df.columns and 'B_renamed' in saved_df.columns
    assert len(saved_df) == 3  # Check if data length is correct
